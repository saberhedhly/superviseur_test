from flask import Blueprint, request, flash, redirect, url_for, render_template, send_file
from database import get_db_connection
from routes.auth import login_required
import io
import openpyxl
from openpyxl.utils import get_column_letter
import sqlite3

instruments_bp = Blueprint('instruments', __name__)

@instruments_bp.route('/add_instrument', methods=['POST'])
@login_required
def add_instrument():
    data = request.form
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO instruments (instrument_name, designation, date_etalonnage, frequence, date_prochain_etalonnage, machine_id)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (data.get('instrument_name'), data.get('designation'), data.get('date_etalonnage'),
          data.get('frequence'), data.get('date_prochain_etalonnage'), data.get('machine_id')))
    conn.commit()
    conn.close()
    flash("✅ Instrument ajouté", "success")
    return redirect(url_for('dashboard.dashboard'))

@instruments_bp.route('/edit_instrument/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_instrument(id):
    conn = get_db_connection()
    cur = conn.cursor()
    
    if request.method == 'POST':
        data = request.form
        cur.execute("""
            UPDATE instruments SET instrument_name=?, designation=?, date_etalonnage=?, frequence=?, date_prochain_etalonnage=?, machine_id=?
            WHERE id=?
        """, (data.get('instrument_name'), data.get('designation'), data.get('date_etalonnage'),
              data.get('frequence'), data.get('date_prochain_etalonnage'), data.get('machine_id'), id))
        conn.commit()
        conn.close()
        flash("✅ Instrument modifié", "success")
        return redirect(url_for('dashboard.dashboard'))
    
    cur.execute("SELECT * FROM instruments WHERE id=?", (id,))
    instrument = cur.fetchone()
    cur.execute("SELECT * FROM machines")
    machines = cur.fetchall()
    conn.close()
    return render_template('edit_instrument.html', instrument=instrument, machines=machines)

@instruments_bp.route('/delete_instrument/<int:id>', methods=['POST'])
@login_required
def delete_instrument(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM instruments WHERE id=?", (id,))
    conn.commit()
    conn.close()
    flash("✅ Instrument supprimé", "success")
    return redirect(url_for('dashboard.dashboard'))

@instruments_bp.route('/export')
@login_required
def export():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''
        SELECT i.id, i.instrument_name, i.designation, i.date_etalonnage, i.frequence, 
               i.date_prochain_etalonnage, i.machine_id, m.machine_name, m.localisation
        FROM instruments i
        LEFT JOIN machines m ON i.machine_id = m.id
    ''')
    instruments = cur.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Instruments"

    headers = ['id', 'instrument_name', 'designation', 'date_etalonnage', 'frequence', 
               'date_prochain_etalonnage', 'machine_id', 'machine_name', 'localisation']
    ws.append(headers)

    for row in instruments:
        ws.append([
            row['id'],
            row['instrument_name'],
            row['designation'],
            row['date_etalonnage'],
            row['frequence'],
            row['date_prochain_etalonnage'],
            row['machine_id'],
            row['machine_name'],
            row['localisation']
        ])

    for col_num, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        adjusted_width = max_length + 2
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = adjusted_width

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name='instruments.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
              )
@instruments_bp.route('/detections')
@login_required
def detections():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT d.*, m.machine_name 
        FROM detections d
        LEFT JOIN machines m ON d.machine_id = m.id
        ORDER BY d.timestamp DESC
        LIMIT 100
    """)
    detections = cur.fetchall()
    conn.close()
    return render_template('detections.html', detections=detections)